import os
import random
from datetime import datetime, timedelta

from faker import Faker
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

fake = Faker('zh_CN')
# TODO 修改为相同用户多次购买

def simulate():
    wb = Workbook()
    ws = wb.active

    headers = ['姓名', '性别', '年龄', '测量时间', '收缩压', '舒张压', '脉压差', '心率', '身高', '体重']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header

    users = {}
    count = random.randrange(500, 2000)
    # generate 1000 rows of fake data
    act_row = 1
    for row in range(2, count):
        # 生成随机日期和时间
        start_date = datetime(2023, 3, 1)
        end_date = datetime(2023, 4, 1)
        random_date = start_date + timedelta(days=random.randint(0, (end_date - start_date).days))
        random_date = random_date.replace(hour=random.randint(0, 23), minute=random.randint(0, 59), second=random.randint(0, 59))

        # 生成时间列表
        num_intervals = random.randint(50, 100)
        time_list = [random_date]
        for _ in range(num_intervals - 1):
            random_date += timedelta(minutes=30)
            time_list.append(random_date)

        name = fake.name()
        if name in users:
            continue
        else:
            users[name] = True
        gender = fake.random_element(elements=('男', '女'))
        age = random.randint(17, 90)
        weight = random.randint(155, 190)
        height = random.randint(40, 110)
        for t in time_list:
            act_row += 1
            d = random.randint(40, 150)
            diff = random.randint(20, 140)
            data = [
                name,
                gender,
                age,
                t.strftime("%Y-%m-%d %H:%M:%S"),
                d + diff,
                d,
                diff,
                random.randint(40, 130),
                weight,
                height
            ]
            for col_num, value in enumerate(data, 1):
                col_letter = get_column_letter(col_num)
                ws[f'{col_letter}{act_row}'] = value

    wb.save('data/simulate/%s.xlsx' % datetime.now().strftime('%Y-%m-%d-%H-%M-%S'))

if __name__ == '__main__':
    simulate()
