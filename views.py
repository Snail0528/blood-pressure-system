import os
import time
import json
from collections import OrderedDict
from datetime import datetime

import numpy as np
import pdfkit
from openpyxl import load_workbook
from flask import Blueprint, render_template, request, jsonify, redirect, url_for, make_response
from flask_login import login_user, logout_user, login_required, current_user, LoginManager
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash

from models import db, Users, PatientImport, PatientData, UserAnalysis
from admin_model_views import PatientImportView
from utils.analysis import analysis
from utils.cluster import cluster_patient
from utils.predict import predict_trends

upload_folder = os.path.join(os.path.dirname(__file__), 'data', 'upload')

bp = Blueprint('login', __name__, url_prefix='')
login_manager = LoginManager()
order_status_map = {
    '0': '支付中',
    '1': '取消支付',
    '2': '支付完成',
}

@bp.route('/')
def index():
    if not current_user.is_authenticated:
        return redirect('/login')
    else:
        return render_template('index.html', user=current_user)

@bp.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        remember = True if request.form.get('remember') == 'on' else False

        user = Users.query.filter_by(username=username).first()

        if not user or not user.check_password(password):
            return jsonify({'status': 'error', 'msg': '用户名或密码错误！'})

        login_user(user, remember=remember, force=True)

        return jsonify({'status': 'success', 'msg': '登录成功！'})

    return render_template('login.html')

@bp.route('/logout')
@login_required
def logout():
    logout_user()
    # 跳转让前端来做，返回数据让前端跳转弹框
    return jsonify({'status': 'success', 'msg': '退出登录成功！'})

@bp.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')

        if password != confirm_password:
            return jsonify({'status': 'error', 'msg': '两次密码输入不一致！'})

        # 查询用户是否已存在
        user = Users.query.filter_by(username=username).first()
        if user:
            return jsonify({'status': 'error', 'msg': '用户名已被占用！'})

        # 创建新用户
        new_user = Users(
            username=username, 
            password=generate_password_hash(password),
            email=email,
            is_admin='0'
        )
        db.session.add(new_user)
        db.session.commit()

        return jsonify({'status': 'success', 'msg': '注册成功，请登录！'})
    
    return render_template('register.html')

@login_manager.user_loader
def load_user(user_id):
    return Users.query.filter_by(username=user_id).first()

@bp.route('/patient')
@login_required
def patient():
    return render_template('patient.html', user=current_user)

@bp.route('/upload', methods=['POST'])
@login_required
def upload_file():
    if 'file' not in request.files:
        return jsonify({'status': 'error', 'msg': '没有选择任何文件！'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'status': 'error', 'msg': '没有选择任何文件！'})
    
    filename = secure_filename(file.filename)
    filepath = os.path.join(upload_folder, filename)
    file.save(filepath)

    wb = load_workbook(filepath)
    ws = wb.active
    rows = ws.iter_rows(min_row=2, values_only=True)

    user_id = current_user.id
    username = current_user.username

    timestamp = int(time.time()) # 生成时间戳
    records = []
    data_records = []

    users = set()
    for row in rows:
        # 解析每一行数据
        export_id = f"{timestamp}"
        patient_name = row[0]
        gender = row[1]
        age = row[2]
        ctime = row[3]
        systolic = row[4]
        diastolic = row[5]
        heartrate = row[7]
        height = row[8]
        weight = row[9]

        analysis_id = '%s_%s' % (user_id, export_id)

        # 创建数据记录并添加到列表中
        if patient_name not in users:
            record = PatientImport(export_id=export_id, patient_name=patient_name, gender=gender, age=age,
                                   height=height, weight=weight, username=username)
            records.append(record)
            users.add(patient_name)
        data_record = PatientData(systolic=systolic, diastolic=diastolic, heartrate=heartrate, ctime=ctime, patient_name=patient_name, export_id=export_id)
        data_records.append(data_record)

    # 将数据记录批量插入到数据库中
    db.session.bulk_save_objects(records)
    db.session.commit()
    db.session.bulk_save_objects(data_records)
    db.session.commit()

    analysis_data = analysis(records, data_records, export_id)
    rfm_data, rfm_result = cluster_patient(filepath)
    new_rfm_data = {}
    rfm_result = rfm_result.sort_values('收缩压', ascending=True)
    cluster_map = OrderedDict({rfm_result.iat[0, 0]: '正常', rfm_result.iat[2, 0]: '低血压患者', rfm_result.iat[1, 0]: '高血压患者'})
    new_rfm_result = []
    sum_count = rfm_result['姓名'].sum()
    for k, v in cluster_map.items():
        new = []
        old = rfm_result[rfm_result['cluster'] == k].iloc[0].tolist()
        for x in old:
            if isinstance(x, float):
                new.append(round(x))
            else:
                new.append(x)
        new.append(round(old[-1] / sum_count, 2))
        new_rfm_result.append((v, new))
        
    for index, x in enumerate(rfm_data['cluster']):
        new = []
        for x in rfm_data.iloc[index].tolist():
            if isinstance(x, (np.int64, np.int32)):
                new.append(int(x))
            elif isinstance(x, np.float64):
                new.append(round(x))
            else:
                new.append(x)
        new_rfm_data.setdefault(cluster_map[x], []).append(new)
    
    count = 0
    for cluster, clu_data in new_rfm_data.items():
        for c in clu_data:
            patient_name = c[0]
            patient_recs = PatientImport.query.filter_by(export_id=export_id, username=current_user.username, patient_name=patient_name).all() 
            for rec in patient_recs:
                rec.cluster = cluster
                db.session.add(rec)
                count += 1
        if count >= 100:
            db.session.commit()
            count = 0
    if count > 0:
        db.session.commit()

    json_data = json.dumps({'analysis_data': analysis_data, 'rfm_data': json.dumps(new_rfm_data), 'rfm_result': new_rfm_result})

    user_analysis = UserAnalysis(analysis_id=analysis_id, analysis_data=json_data, username=username, export_id=export_id)
    db.session.add(user_analysis)
    db.session.commit()

    return jsonify({'status': 'success', 'msg': '上传成功，已生成分析！'})

@bp.route('/get_analysis_history')
@login_required
def get_analysis_history():
    results = UserAnalysis.query.filter_by(username=current_user.username).all()
    return sorted([x.export_id for x in results], reverse=True)

@bp.route('/get_analysis_data', methods=['GET', 'POST'])
@login_required
def get_analysis_data():
    export_id = request.json.get('export_id')
    result = UserAnalysis.query.filter_by(username=current_user.username, export_id=export_id).first()
    if not result or not result.analysis_data:
        return {'status': 'warning', 'msg': '没有查到分析数据，请联系管理员！'}
    result = json.loads(result.analysis_data)
    if not isinstance(result['rfm_data'], dict):
        result['rfm_data'] = json.loads(result['rfm_data'])
    new_result = {}
    for key, value in result['rfm_data'].items():
        new_value = []
        for elem in value:
            new_elem = elem[1:] + elem[:1]
            new_value.append(new_elem)
        new_result[key] = new_value
    result['rfm_data'] = new_result
    return {'status': 'success', 'msg': '获取分析数据成功！', 'data': result}

@bp.route('/get_patient_data', methods=['GET', 'POST'])
@login_required
def get_patient_data():
    export_id = request.json.get('export_id')
    results = PatientImport.query.filter_by(username=current_user.username, export_id=export_id).all()
    if not results:
        return {'status': 'warning', 'msg': '没有查到分析数据，请联系管理员！'}
    # fields = ['patient_name', 'gender', 'age', 'height', 'weight', 'cate', 'cluster']
    fields = ['patient_name', 'gender', 'age', 'height', 'weight', 'cate']
    new_results =  []
    for r in results:
        new_results.append([r.__getattribute__(k) for k in fields])
    results = {
        'labels': [PatientImportView.column_labels[k] for k in fields],
        'data': new_results
    }
    return {'status': 'success', 'msg': '获取分析数据成功！', 'data': results}

@bp.route('/predict', methods=['GET', 'POST'])
@login_required
def predict():
    if request.method == 'GET':
        export_time =  request.values['export_time']
        patient_name =  request.values['patient_name']
        export_id = int(datetime.strptime(export_time, '%Y-%m-%d %H:%M:%S').timestamp())
        username = current_user.username
        patient_data, old_data, predict_data = predict_trends(export_id, username, patient_name)
        patient_data = PatientImport.query.filter_by(export_id=export_id, username=username, patient_name=patient_name).first()
        patient_data = {k: v for k, v in patient_data.__dict__.items() if not k.startswith('_')}
        predict_data.index = predict_data.index.strftime('%Y-%m-%d %H:%M:%S')
        old_data.index = old_data.index.strftime('%Y-%m-%d %H:%M:%S')
        predict_data = {
            'labels': predict_data.index.to_list(),
            'data': [
                [int(x) for x in predict_data['Systolic'].to_list()],
                [int(x) for x in predict_data['Diastolic'].to_list()],
                [int(x) for x in predict_data['HeartRate'].to_list()],
            ]
        }
        old_data = {
            'labels': old_data.index.to_list(),
            'data': [
                [int(x) for x in old_data['systolic'].to_list()],
                [int(x) for x in old_data['diastolic'].to_list()],
                [int(x) for x in old_data['heartrate'].to_list()],
            ]
        }
        return render_template('predict.html', user=current_user, old_data=old_data, predict_data=predict_data, patient_data = patient_data)

@bp.route('/profile')
@login_required
def profile():
    return render_template('profile.html', user=current_user)

@bp.route('/analysis')
@login_required
def analysis_interface():
    return render_template('analysis.html', user=current_user)
